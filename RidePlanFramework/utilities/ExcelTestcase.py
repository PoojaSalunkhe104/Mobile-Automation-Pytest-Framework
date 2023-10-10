import openpyxl

ExcelFilepath = "/Users/macbook/PycharmProjects/RidePlanFramework/TestData/Data.xlsx"

class ReadExcelFile:
    # to open the workbook
    # workbook object is created

    @staticmethod
    def readExcelData(self, TestCaseName):
        wb_obj = openpyxl.load_workbook(ExcelFilepath)
        sheet_obj = wb_obj.active

        for row in range(1, sheet_obj.max_row + 1):
            Value = sheet_obj.cell(row=row, column=1).value
            if Value == TestCaseName.strip():
                for col in range(1, sheet_obj.max_column + 1):
                    RowHeaderNames = sheet_obj.cell(row=1, column=col).value
                    match RowHeaderNames:
                        case 'Address':
                            Address = sheet_obj.cell(row=row, column=col).value
                        case 'Name':
                            Name = sheet_obj.cell(row=row, column=col).value

                return Address, Name;






